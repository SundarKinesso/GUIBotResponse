import os
import openpyxl
import win32com.client as win32
import win32api
import time
import ctypes

# Open the workbook
src_wb = openpyxl.load_workbook('Sizmek_TS.xlsx')
sh1 = src_wb['1074464166']
tgt_wb = openpyxl.load_workbook('RUF_US.xlsx')
sh2 = tgt_wb["Test_sheet"]

# Select the source sheet and target sheet
src_sheet = sh1
tgt_sheet = sh2

# Copy the data in column H, rows 24 to 30
for row in range(24, 30):
    value = src_sheet.cell(row=row, column=8).value
    tgt_sheet.cell(row=row-13, column=5).value = value
    tgt_wb.save("RUF_US.xlsx")

# Save the workbook
src_wb.save("Sizmek_TS.xlsx")
tgt_wb.save("RUF_US.xlsx")

# Open the Excel application
excel = win32.gencache.EnsureDispatch('Excel.Application')

# Make the application visible
excel.Visible = True

# Open the workbook
wb = excel.Workbooks.Open(os.path.abspath('RUF_US.xlsx'))
# Select the RUF_US sheet
sh = wb.Sheets("Test_sheet")
sh.Activate()

time.sleep(0.1)
ctypes.windll.user32.MessageBoxW(0, "Copy paste is done", "Message", 0)