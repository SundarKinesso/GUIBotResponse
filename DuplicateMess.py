import os
import openpyxl
import win32com.client as win32
import win32api
import time
import ctypes

# Open the RPA_T1 workbook
wb2 = openpyxl.load_workbook('Sizmek_TS.xlsx')
sh1 = wb2['1074464166']

sh2 = wb2.copy_worksheet(sh1)

# Give the new sheet a different name
sh2.title = "Duplicate of 1074464166"

# Clear the data from the newly duplicated sheet
for row in sh2.iter_rows():
    for cell in row:
        cell.value = None
# Copy the contents of H24:H29 from the original sheet
H24_H200 = sh1['H24:H200']

# Paste the contents to the duplicate sheet
for row in H24_H200:
    for cell in row:
        sh2[cell.coordinate].value = cell.value
        time.sleep(0.1)

# Set the value of the first cell in the first row of the new sheet to "Start Date"
sh2.cell(row=1, column=8).value = "Start Date"

# Save the workbook with the new sheet
wb2.save('Sizmek_TS.xlsx')

# Open the Excel application
excel = win32.gencache.EnsureDispatch('Excel.Application')

# Make the application visible
excel.Visible = True

# Open the workbook
wb = excel.Workbooks.Open(os.path.abspath('Sizmek_TS.xlsx'))

# Select the new sheet
sh = wb.Sheets("Duplicate of 1074464166")
sh.Activate()

time.sleep(0.1)
ctypes.windll.user32.MessageBoxW(0, "Duplicate of 1074464166 is done", "Message", 0)