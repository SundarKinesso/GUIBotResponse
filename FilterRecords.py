from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Load the workbook
wb = load_workbook("Sizmek_TS.xlsx")

# Select the active sheet
ws = wb.active

# Define the column to filter
column_to_filter = "J"

# Define the text filter
text_filter = "Banner"

# Create a data validation object
dv = DataValidation(type="custom", formula1=f"=NOT(ISERROR(SEARCH(\"{text_filter}\",{column_to_filter}1)))")

# Set the range of cells to apply the validation
dv.ranges = f"{column_to_filter}1:{column_to_filter}{ws.max_row}"

# Add the validation to the worksheet
ws.add_data_validation(dv)

# Save the workbook
wb.save("filtered_records.xlsx")