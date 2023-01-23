# importing packages
import logging
logger = logging.getLogger()
logger.setLevel(logging.CRITICAL)
from chatterbot import ChatBot
from chatterbot.trainers import ChatterBotCorpusTrainer
from tkinter import *
import tkinter as tk
import openpyxl
import tkinter.messagebox as tmsg
import time
from datetime import datetime
import re


# create chatbot instance
# name - name of Python chatbot
my_bot = ChatBot(name = "ChatterBot", storage_adapter = "chatterbot.storage.SQLStorageAdapter")

# training the chatbot
trainer = ChatterBotCorpusTrainer(my_bot)
trainer.train("chatterbot.corpus.english.greetings")
preprocessors = ['chatbot.preprocessors.convert_to_ascii']
#Open the RPA_T1 workbook
wb2 = openpyxl.load_workbook('Sizmek_TS.xlsx')
#sheet2 = wb2.active
sh1 = wb2['1074464166']
l1=sh1.iter_rows(min_row=1,max_row=1,max_col=5,values_only=True)
rset=sh1.iter_rows(min_row=24,max_row=28,values_only=True)

l1=[r for r in l1] # Prepare list for column headers
rset=[r for r in rset] # Prepare list with data
#print(l1)
#print(rset)

def exec_time(function):
    def wrapper(*args, **kwargs):
        start_time = time.perf_counter()
        result = function(*args, **kwargs)  # executing function
        end_time = time.perf_counter()

        difference = str((end_time - start_time) * 1000)  # their difference
        print(f"{function.__name__} function executed in : {difference} mil sec")

        return result

    return wrapper

#Reading all dates for Start date
@exec_time
def start_date():
    data1 = sh1['H24':'H687']
    count = 0
    for cell in data1:
        for x in cell:
            if x.value:
                count += 1
    return count
print(f"[Number of records: {start_date()}]")

# define the column and filter values
#column_to_filter = 'J'
#filter_values = ['In-Banner', 'In-Stream video', 'In-Stream Video tracking']

# create a list to store the rows to keep
#rows_to_keep = []

# iterate through the rows and filter the column
#for row in sh1.iter_rows():
#    for cell in row:
#        if cell.column == column_to_filter and cell.value in filter_values:
#            rows_to_keep.append(cell.row)
#            break

# iterate through the rows and delete those that don't match the filter
#for row in sh1.iter_rows():
#    if row[0].row not in rows_to_keep:
#        sh1.delete_rows(row[0].row, 1)

#Checking for Dimensions
@exec_time
def dimensions():
    data1 = sh1['L24':'L29']
    for cell in data1:
        for x in cell:
            print(x.value)
            if x.value == '0x0':
                print('Mismatch found')
            elif x.value == '1x0':
                print('Mismatch found')
            else:print("Checked Dimensions correctly")
print(dimensions())
#Checking for 3rd party
# Iterate through the cells in the range 'C24:29'
for row in range(24, 30):
    for col in range(3, 4):
        cell = sh1.cell(row=row, column=col)
        match = re.search(r'_\d{1}rd', cell.value)
        if match:
            print(match.group())
# Define the regex pattern for the date format
date_pattern = re.compile(r'\d{8}')

# Iterate through the cells in the range "C24:C29"
for row in range(24, 30):
    cell = sh1.cell(row=row, column=3)
    match = date_pattern.search(cell.value)
    if match:
        # Extract the date from the cell value
        date = match.group()
        # Reformat the date to "06/19/2022" format
        date = date[4:6] + '/' + date[6:] + '/' + date[:4]
        # Find the matching date in the "H24:H29" range
        for col in range(8, 9):
            match_cell = sh1.cell(row=row, column=col)
            if match_cell.value == date:
                print(f'Start date {date} found in cell {match_cell.coordinate}')

#Checking for End date from string to column
# Iterate through the range of cells
for c_row, h_row in zip(sh1["C24:C29"], sh1["I24:I29"]):
    for c_cell, h_cell in zip(c_row, h_row):
        # Extract the date string from the cell using regular expressions
        date_string = re.search(r'\d{8}', c_cell.value).group()
        # Convert the date string to a date object
        date_object = datetime.strptime(date_string, "%Y%m%d").date()
        # Match the date object with the date column
        if h_cell.value == date_object.strftime("%m/%d/%Y"):
            print("Match found for End date at cell " + h_cell.coordinate)

#Define the Bot response system
def chat():
    output.delete(1.0, "end")
    text = str(input.get(1.0, "end"))
    response1 = my_bot.get_response(text)
    input.delete(2.0, "end")
    output.insert(2.0, response1)
    input.focus_set()

# creating a tkinter window
workbot_wd = tk.Tk()
workbot_wd.geometry("500x500")
workbot_wd.title("Sizmek workflow Bot")
workbot_wd["bg"] = "#00b1a4"


input_label = Label(workbot_wd, text = "Hi I am your Workflow bot", background = "#FFFFFF")
input_label.grid(row = 0, column = 3, padx = 5, pady = 5)

# taking input
input= tk.Text(workbot_wd, width = 40, height = 2)
input.grid(row = 1, column = 3, padx = 25, pady = 5)

#Takes text by clicking on menu text and populates in input box
def on_menu_item_click(item_text):
    input.delete(1.0, "end")
    input.insert(1.0, item_text)
    input.focus_set()


send = Button(workbot_wd, command = chat, text = "Enter Command")
send.grid(row = 1, column = 4, padx = 5, pady = 5)


#send = Button(chatbot_wd, command = excel_data, text = "SEND")
#send.grid(row = 1, column = 3, padx = 5, pady = 5)

# giving output
output = Text(workbot_wd, width = 40, height = 8)
output.grid(row = 2, column = 1, columnspan = 3, padx = 25, pady = 5)

def QATopic():
    tmsg.showinfo("Media Plan QA")
def Commanlines():
    value=tmsg.askquestion("Upload your TS?")
    if value=="yes":
        msg="Go ahead with commands"
    else :
        msg = "SORRY ... we will try to improve it "
    tmsg.showinfo("Experience",msg)

mainmenu= tk.Menu (workbot_wd)

m1 = tk.Menu (mainmenu,tearoff=0)
m1.add_command(label ="Complete checking all dimensions records now",command = Commanlines)
m1.add_command(label="Upload your TS?", command=lambda: on_menu_item_click("Upload your TS?"))
m1.add_command(label="Filter In-Banner for QA checks?", command=lambda: on_menu_item_click("Filter banner for QA checks"))
m1.add_command(label="Filter In-Stream Video for QA checks?", command=lambda: on_menu_item_click("Filter In-Stream Video for QA checks"))
m1.add_command(label="Filter In-stream Video Tracking for QA checks?", command=lambda: on_menu_item_click("Filter In-Stream Video Tracking for QA checks"))
m1.add_command(label="Check third party for first five records",
               command=lambda: on_menu_item_click("Check third party for first five records"))
m1.add_command(label="Read 687 records dimension and check for 0x0 and 1x0 in columns",
               command=lambda: on_menu_item_click("Read 687 records dimension and check for 0x0 and 1x0 in columns"))
m1.add_command(label="Check Start date for first six records from string to column matching",
               command=lambda: on_menu_item_click("Check Start date for first six records from string to column matching"))
workbot_wd.config(menu=mainmenu)
mainmenu.add_cascade(label="Commands QA Topic-1",menu=m1)
workbot_wd.mainloop()
