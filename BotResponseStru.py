# importing packages
#Make sure for Python 3.9 version the numpy version
#should be numpy==1.22.0
import os
from tkinter import *
import tkinter as tk
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import tkinter.messagebox as tmsg
#import matplotlib.pyplot as plt
import time
from datetime import datetime
import re
#from PIL import ImageTk,Image
import pyttsx3
#Open the RPA_T1 workbook
wb2 = openpyxl.load_workbook('Sizmek_TS.xlsx')
#sheet2 = wb2.active
sh1 = wb2['1074464166']
def exec_time(function):
    def wrapper(*args, **kwargs):
        start_time = time.perf_counter()
        result = function(*args, **kwargs)  # executing function
        end_time = time.perf_counter()

        difference = str((end_time - start_time) * 1000)  # their difference
        print(f"{function.__name__} function executed in : {difference} mil sec")

        return result

    return wrapper
#Checking for Dimensions
@exec_time
def dimensions():
    data1 = sh1['L24':'L29']
    for cell in data1:
        for x in cell:
            print(x.value)
            if x.value == '0x0':
                print('Mismatch found')
            elif x.value == '1x0':
                print('Mismatch found')
            else:print("Checked Dimensions correctly")
#print(dimensions())
#Read all dimensions of L column
def read_l_column():
    data = sh1['L1':'L' + str(sh1.max_row)]
    valuesl = []
    for row in data:
        for cell in row:
            valuesl.append(cell.value)
    return valuesl
valuesdl = read_l_column()
total_value = len(valuesdl)
print("Total value of dimensions column:", total_value)

#Finding mismatch for specific row and column
def mismatch_dimensions():
    data1 = sh1['L24':'L29']
    result = []
    for row in data1:
        for cell in row:
            if cell.value == '0x0':
                print(f"Mismatch found for 0x0 at Row: {cell.row}, Column: {cell.column}")
                result.append(f"\n Mismatch found for 0x0 in Row:\n{cell.row},Column:{cell.column}")
            if cell.value == '1x0':
                print(f"Mismatch found for 1x0 at Row: {cell.row}, Column: {cell.column}")
                result.append(f"\n Mismatch found for 1x0 in Row:\n{cell.row},Column:{cell.column}")
    return result

#Read Dimensions value for In Banner from placement type column

#Checking for 3rd party
# Iterate through the cells in the range 'C24:29'
for row in range(24, 30):
    for col in range(3, 4):
        cell = sh1.cell(row=row, column=col)
        match = re.search(r'_\d{1}rd', cell.value)
        if match:
            print(match.group())
#Extract the Total count of 3rd party records
def extract_third_party_records(sh1):
    third_party_records = []
    for row in range(24, 687):
        for col in range(3, 4):
            cell = sh1.cell(row=row, column=col)
            match = re.search(r'_\d{1}rd', cell.value)
            if match:
                third_party_records.append(cell.value)
    return len(third_party_records)
total_value = extract_third_party_records(sh1)
print("Total third party records:", total_value)

#Output all non third party records
def extract_non_third_party_records(sh1):
    non_third_party_records = []
    for row in range(24, 687):
        for col in range(3, 4):
            cell = sh1.cell(row=row, column=col)
            match = re.search(r'_\d{1}rd', cell.value)
            if not match:
                non_third_party_records.append(cell.value)
    return len(non_third_party_records)

non_third_party_records = extract_non_third_party_records(sh1)
print("Non-third party records:", non_third_party_records)
#Plot third and non third party records

#plot_third_party_records(sh1)
#Reading all dates for Start date
@exec_time
def start_date():
    data1 = sh1['H24':'H687']
    count = 0
    for cell in data1:
        for x in cell:
            if x.value:
                count += 1
    return count
print(f"[Number of records: {start_date()}]")

#Checking for End date from string to column
# Iterate through the range of cells
#Matching with start date extracted with H column
#Matching with start date extracted with H column
def extract_date(string):
    date_string = re.search(r'\d{8}', string).group()
    date_object = datetime.strptime(date_string, "%Y%m%d")
    formatted_date = date_object.strftime("%m/%d/%Y")
    print("Start Date extracted: ",formatted_date)
    return formatted_date
def extract_start_date_matching(formatted_date, date_range,sh1):
    match_found = False
    for row in range(date_range[0][1], date_range[1][1] + 1):
        column = ord(date_range[0][0]) - ord('A') + 1
        cell = sh1.cell(row=row, column=column)
        if formatted_date != cell.value:
            continue
        else:
            if not match_found:
                match_found = True
            print(f"Matching start Date extracted in cell {cell.coordinate}: {formatted_date}")
    if not match_found:
        print(f"No matching start date found in the range "
              f"{date_range[0][0]}{date_range[0][1]} - {date_range[1][0]}{date_range[1][1]}")
    return match_found
#Test it with the code
extract_start_date_matching('04/18/2022',[('H',24),('H',29)],sh1)
#For end dates
def extract_date_from_string(string):
    s = pd.Series([string])
    datestr = s.str.extract(r'(20220717)',expand=False)[0]
    dateobj = datetime.strptime(datestr,'%Y%m%d')
    return dateobj.strftime('%m/%d/%Y')

date_string = "20220418_20220717"
formatted_date = extract_date_from_string(date_string)

def extract_date_from_range(sh1):
    for row in sh1.iter_rows(min_row=24, max_row=29, min_col=3, max_col=3):
        for cell in row:
            if cell.value:
                date_string = cell.value
                formatted_date = extract_date_from_string(date_string)
                print("End Date extracted: ", formatted_date)
extract_date_from_range(sh1)

#Matching with End date extracted with I column
def extract_end_date_matching(string, date_range):
    date_string = re.search(r'\d{8}', string).group()
    date_object = datetime.strptime(date_string, "%Y%m%d")
    formatted_date = date_object.strftime("%m/%d/%Y")
    for row in range(date_range[0][1], date_range[1][1] + 1):
        column = ord(date_range[0][0]) - ord('A') + 1
        cell = sh1.cell(row=row, column=column)
        if formatted_date != cell.value:
            print(f"Mismatch found in cell {cell.coordinate}")
            return f"Mismatch found in cell {cell.coordinate}"
        else:
            print(f"Matching end Date extracted in cell {cell.coordinate}: {formatted_date}")
            return f"\nMatching end Date extracted in cell \n{cell.coordinate}: {formatted_date}"

print(extract_end_date_matching('20220717',[('I',24),('I',29)]))

#Bot Response
def chat():
    import time
    output.delete(1.0, "end")
    text = str(input.get(1.0, "end"))
    if text.strip() == 'Read Dimension Columns':
        values = read_l_column()
        total_value = len(values)
        for char in f'Total value of dimensions column:\n{total_value}':
            output.insert("end", char)
            output.update()
            time.sleep(0.1)
    elif text.strip() == 'Check mismatch in dimensions':
        result = mismatch_dimensions()
        if result:
            for item in result:
                output.insert(1.0, item)
                input.focus_set()
    elif text.strip() == 'Count total third party records':
        count = extract_third_party_records(sh1)
        total_count = int(count)
        output.delete(1.0, "end")
        output.insert(1.0, f'Total third party records:\n{total_count}')
        engine = pyttsx3.init()
        engine.say(f'Total third party records are {total_count}')
        engine.runAndWait()
        input.focus_set()
    elif text.strip() == 'Count total non third party records':
        non_third_count = extract_non_third_party_records(sh1)
        total_non_third_count = int(non_third_count)
        output.delete(1.0, "end")
        output.insert(1.0, f'Total non third party records:\n{total_non_third_count}')
        engine = pyttsx3.init()
        engine.say(f'Total non third party records:\n{total_non_third_count}')
        engine.runAndWait()
        input.focus_set()
    elif text.strip() == 'Count start date records':
        num_records = start_date()
        output.delete(1.0, "end")
        output.insert(1.0, f'Number of start date records:\n{num_records}')
        engine = pyttsx3.init()
        engine.say(f'Number of start date records:\n{num_records}')
        engine.runAndWait()
        input.focus_set()

    elif text.strip() == 'Count end date records':
        num1_records = start_date()
        output.delete(1.0, "end")
        output.insert(1.0, f'Number of end date records:\n{num1_records}')
        input.focus_set()
    elif text.strip().startswith("Extracted start date validated with Start date column"):
        match_found = extract_start_date_matching('04/18/2022', [('H', 24), ('H', 29)],sh1)
        output.delete(1.0, "end")
        if match_found:
            output.insert(1.0, f'Extracted start date validated in range:{"H24:H29"}:\n{match_found}')
        else:
            output.insert(1.0, f'No matching start date found')
        input.focus_set()
    elif text.strip().startswith("Extracted end date validated with End date column"):
        string1_date_range = extract_end_date_matching('20220717',[('I',24),('I',29)])
        output.delete(1.0, "end")
        if string1_date_range:
            output.insert(1.0, f'Extracted end date validated in range:\n{"I24:I29"}')
        else:
            output.insert(1.0, f'No matching end date found')
    elif text.strip().startswith("Copy paste from Sizmek_TS to RUF_US"):
        output.delete(1.0, "end")
        import time
        import ctypes
        import win32com.client as win32
        # Open the workbook
        src_wb = openpyxl.load_workbook('Sizmek_TS.xlsx')
        shcp1 = src_wb['1074464166']
        tgt_wb = openpyxl.load_workbook('RUF_US.xlsx')
        shcp2 = tgt_wb["Test_sheet"]

        # Select the source sheet and target sheet
        src_sheet = shcp1
        tgt_sheet = shcp2

        # Copy the data in column H, rows 24 to 30
        for row in range(24, 30):
            value = src_sheet.cell(row=row, column=8).value
            tgt_sheet.cell(row=row - 13, column=5).value = value
            tgt_wb.save("RUF_US.xlsx")

        # Save the workbook
        src_wb.save("Sizmek_TS.xlsx")
        tgt_wb.save("RUF_US.xlsx")

        # Open the Excel application
        excel = win32.gencache.EnsureDispatch('Excel.Application')

        # Make the application visible
        excel.Visible = True

        # Open the workbook
        wb = excel.Workbooks.Open(os.path.abspath('RUF_US.xlsx'))
        # Select the RUF_US sheet
        sh = wb.Sheets("Test_sheet")
        sh.Activate()

        time.sleep(0.1)
        ctypes.windll.user32.MessageBoxW(0, "Copy paste is done", "Message", 0)
        output.insert(1.0, f'Check the message once copy paste\n '
                           f'is done')
        input.focus_set()
    else:
        output.delete(1.0, "end")
        output.insert(1.0, 'Hello! How can I help you today?')
    input.focus_set()

def on_menu_item_click(item_text):
    input.delete(1.0, "end")
    input.insert(1.0, item_text)
    input.focus_set()

# creating a tkinter window
workbot_wd = tk.Tk()
workbot_wd.geometry("500x500")
workbot_wd.title("Sizmek Workflow Bot")
workbot_wd["bg"] = "#00b1a4"


input_label = Label(workbot_wd, text = "Hi I am your Workflow bot", background = "#FFFFFF")
input_label.grid(row = 0, column = 3, padx = 5, pady = 5)

# taking input
input= tk.Text(workbot_wd, width = 40, height = 2)
input.grid(row = 1, column = 3, padx = 25, pady = 5)

send = Button(workbot_wd, command = chat, text = "Enter Command")
send.grid(row = 1, column = 4, padx = 5, pady = 5)

# giving output
output = Text(workbot_wd, width = 40, height = 8)
output.grid(row = 2, column = 1, columnspan = 3, padx = 25, pady = 5)

mainmenu= tk.Menu (workbot_wd)
m1 = tk.Menu (mainmenu,tearoff=0)
m1.add_command(label="Read Dimension Columns", command=lambda: on_menu_item_click("Read Dimension Columns"))
m1.add_command(label="Check mismatch in dimensions", command=lambda: on_menu_item_click("Check mismatch in dimensions"))
m1.add_command(label="Count total third party records", command=lambda: on_menu_item_click("Count total third party records"))
m1.add_command(label="Count total non third party records", command=lambda: on_menu_item_click("Count total non third party records"))
#m1.add_command(label="Plot Third Party Records", command=lambda: on_menu_item_click("Plot Third Party Records"))
m1.add_command(label="Count start date records", command=lambda: on_menu_item_click("Count start date records"))
m1.add_command(label="Extracted start date validated with Start date column", command=lambda: on_menu_item_click("Extracted start date validated with Start date column"))
m1.add_command(label="Count end date records", command=lambda: on_menu_item_click("Count end date records"))
m1.add_command(label="Extracted end date validated with End date column", command=lambda: on_menu_item_click("Extracted end date validated with End date column"))
m1.add_command(label="Copy paste from Sizmek_TS to RUF_US", command=lambda: on_menu_item_click("Copy paste from Sizmek_TS to RUF_US"))
workbot_wd.config(menu=mainmenu)
mainmenu.add_cascade(label="Commands QA Topic-1",menu=m1)
workbot_wd.mainloop()
